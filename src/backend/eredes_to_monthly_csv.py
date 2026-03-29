from __future__ import annotations

import argparse
import csv
from collections import defaultdict
from calendar import monthrange
from datetime import datetime, time
from pathlib import Path
from zoneinfo import ZoneInfo

from openpyxl import load_workbook


LISBON = ZoneInfo("Europe/Lisbon")

# Limites plausiveis de consumo mensal (kWh)
# 30 kWh = minimo conservador (casa desocupada)
# 5000 kWh = maximo conservador (consumo industrial nao esperado em residencial)
MIN_MONTHLY_KWH = 30
MAX_MONTHLY_KWH = 5000


def is_daily_cycle_vazio(local_dt: datetime) -> bool:
    is_dst = bool(local_dt.dst())
    current_time = local_dt.time()
    if is_dst:
        return current_time >= time(23, 0) or current_time < time(9, 0)
    return current_time >= time(22, 0) or current_time < time(8, 0)


def extract_interval_kwh(row: tuple) -> float | None:
    registered_power_kw = row[7]
    measured_power_kw = row[3]
    power_kw = registered_power_kw if registered_power_kw is not None else measured_power_kw
    if power_kw is None:
        return None
    return float(power_kw) * 0.25


def pick_sheet(workbook) -> Any:
    for candidate in ("Leituras", "Consumos"):
        if candidate in workbook.sheetnames:
            return workbook[candidate]
    return workbook[workbook.sheetnames[0]]


def detect_data_start_row(worksheet) -> int:
    for idx, row in enumerate(worksheet.iter_rows(min_row=1, max_row=40, values_only=True), start=1):
        values = [str(cell).strip() if cell is not None else "" for cell in row]
        if len(values) >= 3 and values[0] == "Data" and values[1] == "Hora":
            return idx + 1
        if len(values) >= 4 and values[1] == "Data" and values[2] == "Hora":
            return idx + 1
    raise ValueError("Nao foi possivel localizar a linha de cabecalho com Data/Hora no XLSX da E-REDES.")


def extract_date_time_and_kwh(row: tuple, date_idx: int, time_idx: int) -> tuple[str, str, float | None] | None:
    if len(row) <= max(date_idx, time_idx):
        return None
    row_date = row[date_idx]
    row_time = row[time_idx]
    if not row_date or not row_time:
        return None

    candidates = []
    if len(row) > 7:
        candidates.append(row[7])
    if len(row) > 6:
        candidates.append(row[6])
    if len(row) > 3:
        candidates.append(row[3])
    if len(row) > 2:
        candidates.append(row[2])

    power_kw = None
    for value in candidates:
        if value in (None, ""):
            continue
        try:
            power_kw = float(value)
            break
        except (TypeError, ValueError):
            continue
    if power_kw is None:
        return None
    return str(row_date), str(row_time), power_kw * 0.25


def is_complete_month(year_month: str, latest_date_by_month: dict[str, datetime.date]) -> bool:
    latest_date = latest_date_by_month[year_month]
    year = int(year_month.split("-")[0])
    month = int(year_month.split("-")[1])
    return latest_date.day == monthrange(year, month)[1]


def convert_xlsx_to_monthly_csv(
    input_path: Path, output_path: Path, drop_partial_last_month: bool = False
) -> Path:
    wb = load_workbook(input_path, data_only=True, read_only=True)
    ws = pick_sheet(wb)
    data_start_row = detect_data_start_row(ws)

    header_row = next(ws.iter_rows(min_row=data_start_row - 1, max_row=data_start_row - 1, values_only=True))
    header_values = [str(cell).strip() if cell is not None else "" for cell in header_row]
    if len(header_values) >= 3 and header_values[0] == "Data" and header_values[1] == "Hora":
        date_idx, time_idx = 0, 1
    else:
        date_idx, time_idx = 1, 2

    monthly: dict[str, dict[str, float]] = defaultdict(
        lambda: {"total_kwh": 0.0, "vazio_kwh": 0.0, "fora_vazio_kwh": 0.0}
    )
    latest_date_by_month: dict[str, datetime.date] = {}

    for row in ws.iter_rows(min_row=data_start_row, values_only=True):
        extracted = extract_date_time_and_kwh(row, date_idx=date_idx, time_idx=time_idx)
        if extracted is None:
            continue
        row_date, row_time, interval_kwh = extracted

        local_dt = datetime.strptime(f"{row_date} {row_time}", "%Y/%m/%d %H:%M").replace(tzinfo=LISBON)
        year_month = local_dt.strftime("%Y-%m")
        current_date = local_dt.date()
        monthly[year_month]["total_kwh"] += interval_kwh
        if is_daily_cycle_vazio(local_dt):
            monthly[year_month]["vazio_kwh"] += interval_kwh
        else:
            monthly[year_month]["fora_vazio_kwh"] += interval_kwh
        if year_month not in latest_date_by_month or current_date > latest_date_by_month[year_month]:
            latest_date_by_month[year_month] = current_date

    if drop_partial_last_month and monthly:
        last_year_month = max(monthly)
        if not is_complete_month(last_year_month, latest_date_by_month):
            del monthly[last_year_month]

    # Bounds check: verificar limites plausiveis antes de escrever output (RES-02)
    for ym, totals in monthly.items():
        if not (MIN_MONTHLY_KWH <= totals["total_kwh"] <= MAX_MONTHLY_KWH):
            raise ValueError(
                f"Consumo fora dos limites plausiveis para {ym}: "
                f"{totals['total_kwh']:.1f} kWh (esperado {MIN_MONTHLY_KWH}–{MAX_MONTHLY_KWH} kWh/mes)"
            )

    output_path.parent.mkdir(parents=True, exist_ok=True)
    with output_path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.writer(handle)
        writer.writerow(["year_month", "total_kwh", "vazio_kwh", "fora_vazio_kwh"])
        for year_month in sorted(monthly):
            totals = monthly[year_month]
            writer.writerow(
                [
                    year_month,
                    f"{totals['total_kwh']:.3f}",
                    f"{totals['vazio_kwh']:.3f}",
                    f"{totals['fora_vazio_kwh']:.3f}",
                ]
            )

    return output_path


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Converte export da E-REDES em CSV mensal com vazio e fora de vazio."
    )
    parser.add_argument("--input", required=True, help="Ficheiro XLSX da E-REDES.")
    parser.add_argument("--output", required=True, help="CSV mensal de saida.")
    parser.add_argument(
        "--drop-partial-last-month",
        action="store_true",
        help="Exclui o ultimo mes se o XLSX nao contiver o mes completo.",
    )
    return parser


def main() -> int:
    parser = build_parser()
    args = parser.parse_args()
    out = convert_xlsx_to_monthly_csv(
        Path(args.input),
        Path(args.output),
        drop_partial_last_month=args.drop_partial_last_month,
    )
    print(out)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
